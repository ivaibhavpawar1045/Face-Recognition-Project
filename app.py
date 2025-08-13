import streamlit as st
import cv2
import face_recognition
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook

# Function to recognize faces in the image
def recognize_faces(known_faces_folder):
    known_face_encodings = []
    known_face_names = []

    # Load known faces
    for file in os.listdir(known_faces_folder):
        if file.endswith(".jpg") or file.endswith(".png"):
            image = face_recognition.load_image_file(os.path.join(known_faces_folder, file))
            encoding = face_recognition.face_encodings(image)[0]
            known_face_encodings.append(encoding)
            known_face_names.append(os.path.splitext(file)[0])

    # Capture video from webcam
    cap = cv2.VideoCapture(0)

    attendance_taken = False  # Flag to track if attendance has been taken
    unknown_detected = False  # Flag to track if unknown face is detected

    while not (attendance_taken or unknown_detected):
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to capture video from webcam.")
            break

        # Convert the image from BGR color to RGB color
        rgb_frame = frame[:, :, ::-1]

        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        # Loop through each face found in the frame
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            # Compare faces
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # Check if any known face matches
            if True in matches and not attendance_taken:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]

                # Log attendance to both CSV and Excel files
                log_attendance(name)
                attendance_taken = True
            else:
                # Save image of unknown face
                unknown_faces_folder = "unknown_faces"
                os.makedirs(unknown_faces_folder, exist_ok=True)
                now = datetime.now().strftime("%Y%m%d_%H%M%S")
                cv2.imwrite(os.path.join(unknown_faces_folder, f"unknown_{now}.jpg"), frame)

                # Display message for unknown face
                st.warning("Unknown face detected.")
                unknown_detected = True

            # Draw rectangle around the face and write name
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.5, (0, 0, 255), 1)

        # Display the resulting image
        cv2.imshow('Video', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release webcam and close all OpenCV windows
    cap.release()
    cv2.destroyAllWindows()

    if attendance_taken:
        st.success("Attendance taken successfully.")
    elif unknown_detected:
        st.info("Unknown face detected.")

# Function to log attendance in both CSV and Excel
def log_attendance(name):
    now = datetime.now()
    date_time = now.strftime("%Y-%m-%d %H:%M:%S")
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Log attendance to CSV
    log_to_csv(name, date, time)

    # Log attendance to Excel
    log_to_excel(name, date, time)

# Function to log attendance to CSV
def log_to_csv(name, date, time):
    csv_file = "attendance.csv"
    
    # Check if the CSV file exists
    if not os.path.exists(csv_file):
        # If the file doesn't exist, create it and add the header
        with open(csv_file, mode="w") as file:
            file.write("Name,Date,Time\n")  # Adding header

    # Append the attendance record to the CSV file
    with open(csv_file, mode="a") as file:
        file.write(f"{name},{date},{time}\n")

# Function to log attendance to Excel
def log_to_excel(name, date, time):
    excel_file = "attendance.xlsx"
    
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # If it exists, open the workbook
        wb = load_workbook(excel_file)
        sheet = wb.active
    else:
        # If it doesn't exist, create a new workbook and add header
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Name", "Date", "Time"])  # Adding header row

    # Add a new row with the current date and time
    sheet.append([name, date, time])

    # Save the workbook
    wb.save(excel_file)

# Function to display attendance from Excel file
def display_attendance():
    # Read the Excel file using pandas
    excel_file = "attendance.xlsx"
    if os.path.exists(excel_file):
        df = pd.read_excel(excel_file)
        st.dataframe(df)  # Display the DataFrame as an interactive table
    else:
        st.warning("No attendance data found!")

def main():
    # Add title and design to the page
    st.title("🧑‍💻 Face Recognition Attendance System")
    st.write("This is the Face Recognition Attendance System. Click the button below to start taking attendance.")

    # Add custom CSS for centering buttons and adding an image as background
    st.markdown(
        """
        <style>
        .stApp {
            background-image: url('https://i1.wp.com/tl360.b-cdn.net/wp-content/uploads/2020/10/How-Facial-Recognition-Works.jpg?fit=1200%2C630&ssl=1');
            background-size: cover;
            background-position: center;
            height: 100vh;
        }
        
        .center-button {
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
        }
        
        .stButton>button {
            width: 200px;
            height: 60px;
            font-size: 20px;
            border-radius: 10px;
            background-color: #2C3E50;
            color: white;
            box-shadow: 0 8px 16px rgba(0, 0, 0, 0.2);
        }
        
        .stButton>button:hover {
            background-color: #34495E;
        }
        </style>
        """, unsafe_allow_html=True)

    # Center buttons in the page
    with st.container():
        col1, col2, col3 = st.columns([1, 2, 1])  # Create 3 columns for layout
        with col2:
            # Button to take attendance
            if st.button("📸 Take Attendance", key="attendance_button", help="Click to start taking attendance.", use_container_width=True):
                known_faces_folder = "known_faces"
                recognize_faces(known_faces_folder)

            # Button to view attendance
            if st.button("📋 View Attendance", key="view_button", help="Click to view the attendance data"):
                display_attendance()  # Display the attendance in a box when clicked

if __name__ == "__main__":
    main()
